from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import Workbook
# ── GPP INTEGRATION POINT A ────────────────────────────────────────────────
# When the GPP workbook builder is ready, import gpp_integration here and
# call gpp_integration.build_gpp_workbook(payload, signatures) inside
# build_delivery_note_xlsx() just before the function returns `data`.
#
# Example:
#   import gpp_integration
#   gpp_bytes = gpp_integration.build_gpp_workbook(payload, signatures)
#   # attach as a second sheet, save alongside, or return separately
# ───────────────────────────────────────────────────────────────────────────
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _set_col_widths(ws) -> None:
    widths = {
        1: 28,
        2: 52,
        3: 28,
        4: 40,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_delivery_note_xlsx(
    payload: dict[str, Any],
    signatures: dict[str, dict[str, Any]],
    output_path: Path | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leveringsbon"

    _set_col_widths(ws)

    rows = [
        ("Date", payload.get("date")),
        ("Client address", payload.get("client_address")),
        ("Plant address", payload.get("plant_address")),
        ("Delivery Note No", payload.get("delivery_note_no")),
        ("Address", payload.get("site_address")),
        ("Departure Time", payload.get("departure_time")),
        ("Arrival Time", payload.get("arrival_time")),
        ("Distance covered (km)", payload.get("distance_km")),
        ("Transport company", payload.get("transport_company")),
        ("Nummerplaat", payload.get("license_plate")),
        ("Transport type", payload.get("transport_type")),
        ("Energy Source", payload.get("energy_source")),
        ("Product/ Mixture type", payload.get("product_mixture_type")),
        ("Application", payload.get("application")),
        ("Certificate", payload.get("certificate")),
        ("Declaration of Performance", payload.get("declaration_of_performance")),
        ("Technical Data Sheet", payload.get("technical_data_sheet")),
        ("Mechanical resistance", payload.get("mechanical_resistance")),
        ("Fuel resistance", payload.get("fuel_resistance")),
        ("De-icing resistance", payload.get("deicing_resistance")),
        ("Bitumen–aggregate affinity", payload.get("bitumen_aggregate_affinity")),
        ("Disposal", payload.get("disposal")),
        ("Bruto (kg)", payload.get("bruto_kg")),
        ("Tare Weight (Empty)", payload.get("tare_weight_empty_kg")),
        ("Net total quantity (ton)", payload.get("net_total_quantity_ton")),
    ]

    ws["A1"].value = "Digitale Leveringsbon (Prototype)"
    ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)

    start_row = 3
    for i, (k, v) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    sig_start = start_row + len(rows) + 2
    ws.cell(row=sig_start, column=1, value="Handtekeningen")
    ws.cell(row=sig_start, column=1).font = ws.cell(row=sig_start, column=1).font.copy(bold=True)

    # Column headers for the signature section
    hdr_row = sig_start + 1
    ws.cell(row=hdr_row, column=1, value="Rol")
    ws.cell(row=hdr_row, column=2, value="Naam ondertekenaar")
    ws.cell(row=hdr_row, column=3, value="Ondertekend op (UTC)")
    ws.cell(row=hdr_row, column=4, value="Handtekening")
    for c in [1, 2, 3, 4]:
        ws.cell(row=hdr_row, column=c).font = ws.cell(row=hdr_row, column=c).font.copy(bold=True)

    sig_rows = [
        ("Handtekening opdrachtgever", "client"),
        ("Handtekening COPRO", "copro"),
        ("Handtekening vervoerder", "transporter"),
        ("Handtekening vergunninghouder", "permit_holder"),
    ]

    for idx, (label, role) in enumerate(sig_rows, start=1):
        r = hdr_row + idx
        ws.cell(row=r, column=1, value=label)
        meta = signatures.get(role)
        if meta:
            ws.cell(row=r, column=2, value=meta.get("signer_name") or "")
            ws.cell(row=r, column=3, value=meta.get("signed_at") or "")
            sig_path = meta.get("signature_path")
            if sig_path and Path(sig_path).exists():
                img = XLImage(sig_path)
                img.height = 60
                img.width = 220
                ws.add_image(img, f"D{r}")
                ws.row_dimensions[r].height = 50
        else:
            ws.cell(row=r, column=2, value="")
            ws.cell(row=r, column=3, value="")

    bio = io.BytesIO()
    wb.save(bio)
    data = bio.getvalue()

    # ── GPP INTEGRATION POINT A (call site) ────────────────────────────────
    # Populate the GPP template with the same payload + signatures here.
    # Uncomment and complete once gpp_integration.py is implemented:
    #
    #   import gpp_integration
    #   gpp_bytes = gpp_integration.build_gpp_workbook(payload, signatures)
    #   if output_path is not None:
    #       gpp_path = output_path.with_stem(output_path.stem + "_GPP")
    #       gpp_path.write_bytes(gpp_bytes)
    # ───────────────────────────────────────────────────────────────────────

    if output_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(data)

    return data
