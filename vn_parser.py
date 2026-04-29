"""
vn_parser.py
============
Parse the "Verantwoordingsnota" (VN) Excel workbook — specifically the
"Info voor GPP" sheet — into structured Python data, ready to be fed
to the GPP tool via :mod:`vn_to_gpp`.

The sheet lists three asphalt-plant variants side-by-side
(columns E, G, I for general data and E/G/I + F/H/J for component
origin + transport mode).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Info voor GPP"

# Column letters (1-based when needed)
PLANT_COLS = {
    # plant_index: (general_value_col, transport_mode_col)
    0: ("E", "F"),
    1: ("G", "H"),
    2: ("I", "J"),
}

# Static row anchors on the Info-voor-GPP sheet
ROW_PLANT_HEADER = 2          # plant ID (number)
ROW_DATE         = 4
ROW_MIXTURE_ID   = 5
ROW_MIX_SB250    = 6
ROW_MIX_EN       = 7
ROW_BINDER_PCT   = 8           # total binder content (%)
ROW_BINDER_REPL  = 9           # binder replacement (%)
ROW_PLANT_LOC    = 11
ROW_PLANT_ENERGY = 12          # aardgas / propaan / elektrisch
ROW_PLANT_CAP    = 13          # t/h
ROW_PROD_TEMP    = 15          # textual range (NOT pushed to GPP)
ROW_BINDER_TYPE  = 19          # header row, value row sits at 20 in this template
ROW_BINDER_NAME  = 20
ROW_BINDER_PCT2  = 23
# Composition table: aggregates start at row 25.  Each row carries
#   C: component name
#   D: percentage  (or string for filler IIa marker)
#   E/G/I: origin city (per plant)
#   F/H/J: aanvoer per (mode of transport)  — sometimes empty (= same as previous plant)
ROW_COMPOSITION_FIRST = 25
ROW_COMPOSITION_LAST  = 49     # additives end here
ROW_BIOGENIC_PCT      = 50
# Performance tests
ROW_ITSR    = 52
ROW_PRD     = 53
ROW_STIFF1  = 54
ROW_STIFF2  = 55
ROW_FATIGUE = 56


# ─────────────────────────────────────────────────────────────────────
#  Data classes
# ─────────────────────────────────────────────────────────────────────

@dataclass
class VNComponent:
    """One row from the composition table."""
    row: int                       # source row in the sheet
    name: str                      # column C
    pct: float                     # column D, in percent (0-100)
    origin: str | None             # per-plant column E/G/I
    mode: str | None               # per-plant column F/H/J  (raw NL term)
    extra: str | None = None       # extra metadata cell (e.g. filler IIa)


@dataclass
class VNPlant:
    """All data for a single asphalt plant column."""
    plant_index: int               # 0, 1, 2
    plant_id: str                  # row 2 (e.g. "801")
    date: str | None
    mixture_id: str | None
    mixture_sb250: str | None
    mixture_en: str | None
    total_binder_pct: float | None    # column D row 8 (e.g. 3.85)
    binder_replacement_pct: float | None
    plant_location: str | None
    plant_energy: str | None          # raw NL: aardgas / propaan / elektrisch
    plant_capacity_tph: float | None
    prod_temp_range: str | None       # NOT used (kept in GPP defaults)
    binder_type: str | None
    binder_origin: str | None
    binder_mode: str | None
    binder_pct: float | None
    components: list[VNComponent] = field(default_factory=list)
    biogenic_pct: float | None = None
    # Performance tests (informational only)
    itsr: float | None = None
    prd: float | None = None
    stiffness_e_modulus: float | None = None
    fatigue_eps6: float | None = None

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["components"] = [asdict(c) for c in self.components]
        return d


@dataclass
class VNData:
    """Full parsed VN workbook contents."""
    source_filename: str | None
    plants: list[VNPlant]

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_filename": self.source_filename,
            "plants": [p.to_dict() for p in self.plants],
        }


# ─────────────────────────────────────────────────────────────────────
#  Parser
# ─────────────────────────────────────────────────────────────────────

def _val(ws, row: int, col: str) -> Any:
    v = ws[f"{col}{row}"].value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_pct(v: Any) -> float | None:
    """Normalize a cell to a 0–100 percentage.

    Excel cells formatted as Percent return their underlying fraction
    (e.g. 5.5% → 0.055), while plain-number cells return 5.5. Treat any
    value <= 1.0 as a fraction and scale it; values > 1 are taken as
    already-percent.
    """
    f = _to_float(v)
    if f is None:
        return None
    if -1.0 <= f <= 1.0:
        return f * 100.0
    return f


def parse(source: str | Path | bytes | io.BytesIO,
          source_filename: str | None = None) -> VNData:
    """Parse a VN workbook from a path, bytes, or BytesIO.

    Returns a :class:`VNData` with three :class:`VNPlant` entries.
    """
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=False)
    elif isinstance(source, io.BytesIO):
        wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    else:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
        if source_filename is None:
            source_filename = Path(source).name

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    plants: list[VNPlant] = []
    for idx in range(3):
        gen_col, mode_col = PLANT_COLS[idx]
        plant_id = _to_str(_val(ws, ROW_PLANT_HEADER, gen_col)) or f"plant_{idx + 1}"

        components: list[VNComponent] = []
        for row in range(ROW_COMPOSITION_FIRST, ROW_COMPOSITION_LAST + 1):
            name = _to_str(_val(ws, row, "C"))
            pct_raw = _val(ws, row, "D")
            pct = _to_float(pct_raw)
            extra = pct_raw if (pct is None and pct_raw is not None) else None
            origin = _to_str(_val(ws, row, gen_col))
            mode = _to_str(_val(ws, row, mode_col))

            # Skip rows where component name is missing or there is no
            # meaningful share / metadata.  Filler row uses string in D.
            if not name:
                continue
            # Skip the "Additieven" / "Aggregaten" sub-header rows
            # (column D contains "%", no real component data).
            if name.lower() in {"additieven", "aggregaten"}:
                continue
            # Skip placeholder rows (0 % AND no qualitative metadata in D)
            # AND no operator-entered origin/mode info to preserve.
            if (pct is None or pct == 0) and not extra and not origin and not mode:
                continue
            components.append(VNComponent(
                row=row,
                name=name,
                pct=pct if pct is not None else 0.0,
                origin=origin,
                mode=mode,
                extra=_to_str(extra),
            ))

        plant = VNPlant(
            plant_index=idx,
            plant_id=plant_id,
            date=_to_str(_val(ws, ROW_DATE, gen_col)),
            mixture_id=_to_str(_val(ws, ROW_MIXTURE_ID, gen_col)),
            mixture_sb250=_to_str(_val(ws, ROW_MIX_SB250, gen_col)),
            mixture_en=_to_str(_val(ws, ROW_MIX_EN, gen_col)),
            total_binder_pct=_to_pct(_val(ws, ROW_BINDER_PCT, gen_col)),
            binder_replacement_pct=_to_pct(_val(ws, ROW_BINDER_REPL, gen_col)),
            plant_location=_to_str(_val(ws, ROW_PLANT_LOC, gen_col)),
            plant_energy=_to_str(_val(ws, ROW_PLANT_ENERGY, gen_col)),
            plant_capacity_tph=_to_float(_val(ws, ROW_PLANT_CAP, gen_col)),
            prod_temp_range=_to_str(_val(ws, ROW_PROD_TEMP, gen_col)),
            binder_type=_to_str(_val(ws, ROW_BINDER_NAME - 0, "C")),  # fallback
            binder_origin=_to_str(_val(ws, ROW_BINDER_NAME, gen_col)),
            binder_mode=_to_str(_val(ws, ROW_BINDER_NAME, mode_col)),
            binder_pct=_to_pct(_val(ws, ROW_BINDER_PCT2, gen_col)),
            components=components,
            biogenic_pct=_to_pct(_val(ws, ROW_BIOGENIC_PCT, gen_col)),
            itsr=_to_float(_val(ws, ROW_ITSR, gen_col)),
            prd=_to_float(_val(ws, ROW_PRD, gen_col)),
            stiffness_e_modulus=_to_float(_val(ws, ROW_STIFF1, gen_col)),
            fatigue_eps6=_to_float(_val(ws, ROW_FATIGUE, gen_col)),
        )
        # Fallback: pick up the actual binder name (row 20 col C may be
        # the binder description)
        if not plant.binder_type:
            plant.binder_type = _to_str(_val(ws, ROW_BINDER_NAME, "C"))
        plants.append(plant)

    return VNData(source_filename=source_filename, plants=plants)
