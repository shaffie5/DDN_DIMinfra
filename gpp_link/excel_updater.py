import re
import copy
import math
import shutil
import zipfile
import tempfile

from pathlib import Path
from .config import logger
from lxml import etree as et
from openpyxl import load_workbook
from .file_manager import FileManager
from werkzeug.routing import ValidationError
from typing import Dict, Any, List, Tuple, Optional, Union

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML_PARSER = et.XMLParser(remove_blank_text=False, recover=True)
XL_RELS_PATH = "xl/_rels/workbook.xml.rels"

# Helper regexes
_CELL_REF_RE = re.compile(r"^([A-Za-z]+)(\d+)$")
_PERCENT_UNESCAPED_RE = re.compile(r"(?<!\\)%")  # matches '%' not preceded by backslash


class SharedStringsManager:
    """
    Small helper to centralize sharedStrings.xml handling.
    - Maintains a list of SI elements (deep copies).
    - Preserves original template/root attributes when saving.
    - Exposes: load(xl_dir), get_text(si), get_or_add(text) -> index, save(xl_dir)
    """

    def __init__(self):
        self._si_list: List[et._Element] = []
        self._text_to_index: Dict[str, int] = {}  # Reverse lookup cache
        self._modified: bool = False
        self._template: Optional[et._Element] = None

    # Context manager support for better resource management
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._si_list.clear()
        self._text_to_index.clear()
        return False

    def load(self, xl_dir: Path) -> None:
        """Load shared strings from file and build reverse lookup index."""
        self._si_list.clear()
        self._text_to_index.clear()
        self._template = None
        shared_strings_file = xl_dir / "sharedStrings.xml"

        if not shared_strings_file.exists():
            logger.debug("No sharedStrings.xml found")
            return

        try:
            tree = et.parse(shared_strings_file, XML_PARSER)
            root = tree.getroot()
            self._template = root

            # Gather SI elements in document order and deep-copy them
            si_elements = root.findall(f".//{{{MAIN_NS}}}si")

            for si in si_elements:
                self._si_list.append(copy.deepcopy(si))

            # Build reverse lookup index
            for idx, si in enumerate(self._si_list):
                si_text = self._si_text(si)
                if si_text not in self._text_to_index:
                    self._text_to_index[si_text] = idx

            logger.info(f"Loaded {len(self._si_list)} shared strings")
        except Exception as ex:
            logger.warning(f"Failed parsing sharedStrings.xml: {ex}")
            self._si_list.clear()
            self._text_to_index.clear()
            self._template = None

    @staticmethod
    def _si_text(si: et._Element) -> str:
        """Extract text content from SI element safely."""
        if si is None:
            return ""

        parts: List[str] = []
        try:
            for t in si.findall(f".//{{{MAIN_NS}}}t"):
                if t is not None and t.text is not None:
                    parts.append(t.text)
            return "".join(parts)
        except (AttributeError, TypeError, et.LxmlError) as e:
            logger.warning(f"Failed to extract text from SI element: {e}")
            return ""

    def get_or_add(self, text: Optional[str]) -> int:
        """Get existing string index or add new string."""
        if text is None:
            text = ""

        # Use reverse lookup cache for O(1) access
        if text in self._text_to_index:
            return self._text_to_index[text]

        # Create a simple SI element (namespace-aware)
        si = et.Element(f"{{{MAIN_NS}}}si")
        t = et.SubElement(si, f"{{{MAIN_NS}}}t")

        if text.startswith(" ") or text.endswith(" "):
            t.set(et.QName("http://www.w3.org/XML/1998/namespace", "space"), "preserve")

        t.text = text
        new_index = len(self._si_list)
        self._si_list.append(si)
        self._text_to_index[text] = new_index
        self._modified = True

        return new_index

    def save(self, xl_dir: Path) -> None:
        """Save shared strings to file if modified."""
        if not self._modified:
            return

        shared_strings_file = xl_dir / "sharedStrings.xml"

        # Build root preserving template attributes & nsmap when possible
        if self._template is not None:
            root_tag = self._template.tag
            root_nsmap = self._template.nsmap.copy() if self._template.nsmap else {}
            root_attrs = {k: v for k, v in self._template.items()}
            root_attrs["count"] = str(len(self._si_list))
            root_attrs["uniqueCount"] = str(len(self._si_list))
            root = et.Element(root_tag, **root_attrs, nsmap=root_nsmap)
        else:
            # Fallback
            root = et.Element(f"{{{MAIN_NS}}}sst",
                              count=str(len(self._si_list)),
                              uniqueCount=str(len(self._si_list)),
                              nsmap={None: MAIN_NS})

        for si in self._si_list:
            root.append(copy.deepcopy(si))

        xml_content = et.tostring(root, xml_declaration=True, encoding="UTF-8", standalone="yes")
        shared_strings_file.write_bytes(xml_content)
        logger.info(f"Saved {len(self._si_list)} shared strings")
        self._modified = False


class ExcelUpdater:
    """Locale-aware Excel updater that preserves target cell formats (including Percent)."""

    # Built-in percent ids
    _PERCENT_BUILTIN_IDS = frozenset({9, 10, 27, 37, 38, 39, 40})

    # Configuration constants
    _DEFAULT_SHEET_NAME = "Sheet1"
    _TEMP_WORKBOOK_NAME = "temp.xlsx"
    _UPDATED_WORKBOOK_NAME = "updated.xlsx"
    _CONTENT_DIR_NAME = "content"
    _WORKSHEETS_DIR_NAME = "worksheets"
    _MAX_SHARED_STRINGS = 1000000
    _MAX_SHEETS = 1000
    _TIMEOUT_SECONDS = 300

    # Common XML namespaces
    _NAMESPACES = {
        "main": MAIN_NS,
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    }

    def __init__(self):
        # Instance-local caches/state
        self._nsmap_cache: Dict[int, Dict[str, str]] = {}
        self._style_to_numfmt: Dict[str, str] = {}
        self.shared_strings = SharedStringsManager()

    # -------------------- Namespace helpers -------------------- #
    def get_nsmap_with_main(self, elem: et._Element) -> Dict[str, str]:
        """Get namespace map ensuring main namespace is present."""
        cache_key = id(elem)
        if cache_key in self._nsmap_cache:
            return self._nsmap_cache[cache_key].copy()

        nsmap = elem.nsmap.copy() if elem.nsmap else {}
        nsmap.setdefault("main", MAIN_NS)
        self._nsmap_cache[cache_key] = nsmap.copy()

        return nsmap

    @staticmethod
    def find_ns(elem: et._Element, path: str, nsmap: Dict[str, str]) -> Optional[et._Element]:
        """Find element with namespace support."""
        return elem.find(path, namespaces=nsmap)

    @staticmethod
    def findall_ns(elem: et._Element, path: str, nsmap: Dict[str, str]) -> List[et._Element]:
        """Find all elements with namespace support."""
        return elem.findall(path, namespaces=nsmap)

    # -------------------- Validation Helpers -------------------- #
    def _validate_inputs(self, source: Path, target: Path, mappings: List[Dict[str, str]]) -> None:
        """Validate input files and mappings."""
        if not source.exists():
            raise FileNotFoundError(f"Source file not found: {source}")

        if source == target:
            raise ValueError("Source and target files cannot be the same")

        if not target.exists():
            raise FileNotFoundError(f"Target file not found: {target}")

        if not mappings:
            logger.info("No mappings provided, nothing to update")
            return

        if not FileManager.validate_excel_file(source):
            raise ValidationError(f"Invalid source Excel file: {source}")

        self._validate_limits(mappings)
        self._validate_mapping_structure(mappings)

    @staticmethod
    def _validate_mapping_structure(mappings: List[Dict[str, str]]) -> None:
        """Validate mapping structure."""
        for i, mapping in enumerate(mappings):
            if not isinstance(mapping, dict):
                raise ValueError(f"Mapping at index {i} is not a dictionary")

            required_keys = {"SourceCell", "TargetCell"}
            missing_keys = required_keys - set(mapping.keys())
            if missing_keys:
                raise ValueError(f"Mapping {i} missing required keys: {missing_keys}")

    def _validate_limits(self, mappings: List[Dict[str, str]]) -> None:
        """Validate that we're within reasonable limits."""
        if len(mappings) > self._MAX_SHEETS * 1000:
            logger.warning(f"Large number of mappings: {len(mappings)}")

        if len(mappings) == 0:
            logger.info("No mappings to process")

    @staticmethod
    def _validate_update_results(updates_by_sheet: Dict[Path, List[Tuple[str, Any]]], total_updates: int) -> None:
        """Validate that updates were applied as expected."""
        expected_total = sum(len(updates) for updates in updates_by_sheet.values())
        if total_updates != expected_total:
            logger.warning(f"Update count mismatch: expected {expected_total}, applied {total_updates}")

        if total_updates == 0:
            logger.warning("No cells were updated - this might indicate configuration issues")

    # -------------------- Source extraction -------------------- #
    @staticmethod
    def extract_source_values(source: Path, mappings: List[Dict[str, str]]) -> Dict[Tuple[str, str], Any]:
        """Extract values from source workbook based on mappings."""
        src_values: Dict[Tuple[str, str], Any] = {}
        processed_sheets: Dict[str, Any] = {}

        try:
            wb = load_workbook(source, data_only=True, read_only=True)

            try:
                for row in mappings:
                    source_sheet_name = (row.get("SourceSheet") or wb.sheetnames[0]).strip()
                    target_sheet = (row.get("TargetSheet") or "").strip()
                    source_cell = (row.get("SourceCell") or "").strip()
                    target_cell = (row.get("TargetCell") or "").strip()

                    # Skip invalid mappings early
                    if not source_cell or not target_cell:
                        continue

                    if source_sheet_name not in processed_sheets:
                        if source_sheet_name not in wb.sheetnames:
                            logger.warning(f"Source sheet '{source_sheet_name}' not found in {source}")
                            continue
                        processed_sheets[source_sheet_name] = wb[source_sheet_name]

                    sheet = processed_sheets[source_sheet_name]

                    try:
                        cell_value = sheet[source_cell].value
                        # Better empty string handling
                        if cell_value == "" or (isinstance(cell_value, str) and cell_value.strip() == ""):
                            cell_value = None
                        src_values[(target_sheet, target_cell)] = cell_value
                    except (KeyError, AttributeError) as ex:
                        logger.warning(f"Failed to read cell {source_cell} in sheet {source_sheet_name}: {ex}")
            finally:
                wb.close()
        except Exception as ex:
            logger.error(f"Failed to load source workbook {source}: {ex}")
            raise

        logger.info(f"Extracted {len(src_values)} values from {len(processed_sheets)} sheets")
        return src_values

    # -------------------- Styles parsing -------------------- #
    def _load_styles_from_xl(self, xl_dir: Path) -> None:
        """Load styles.xml and build style_id -> numfmt mapping for percent detection."""
        self._style_to_numfmt.clear()
        styles_file = xl_dir / "styles.xml"

        if not styles_file.exists():
            logger.debug("No styles.xml found, skipping style loading")
            return

        try:
            tree = et.parse(styles_file, XML_PARSER)
            root = tree.getroot()
            nsmap = self.get_nsmap_with_main(root)

            # Parse number formats into a mapping
            numfmt_map: Dict[str, str] = {}
            for nf in ExcelUpdater.findall_ns(root, ".//main:numFmt", nsmap):
                nf_id = nf.attrib.get("numFmtId")
                code = (nf.text or "").strip()

                if nf_id and code:
                    numfmt_map[nf_id] = code

            # Parse cell styles
            xfs = self.findall_ns(root, ".//main:cellXfs/main:xf", nsmap)
            for idx, xf in enumerate(xfs):
                num_fmt_id = xf.attrib.get("numFmtId")
                fmt = ""

                if num_fmt_id:
                    fmt = numfmt_map.get(num_fmt_id, "")
                    if not fmt:
                        fmt = f"builtin:{num_fmt_id}"

                self._style_to_numfmt[str(idx)] = fmt

            logger.debug(f"Loaded {len(self._style_to_numfmt)} style mappings")
        except Exception as ex:
            logger.warning(f"Failed parsing styles.xml: {ex}")
            self._style_to_numfmt.clear()

    def _style_id_is_percent(self, style_id: Optional[str]) -> bool:
        """Check if style ID represents a percentage format."""
        if not style_id:
            return False

        fmt = self._style_to_numfmt.get(str(style_id), "")
        if fmt and _PERCENT_UNESCAPED_RE.search(fmt):
            return True

        # Use constant set for faster lookup
        if fmt.startswith("builtin:"):
            try:
                built_id = int(fmt.split(":", 1)[1])
                return built_id in self._PERCENT_BUILTIN_IDS
            except (ValueError, IndexError):
                pass

        if "%" in fmt:
            return True

        return False

    # -------------------- Cell reference helpers -------------------- #
    @staticmethod
    def _parse_cell_ref(cell_ref: str) -> Tuple[str, str]:
        """Parse cell reference into column letters and row index."""
        match = _CELL_REF_RE.match(cell_ref)
        if not match:
            raise ValueError(f"Invalid cell reference: {cell_ref}")

        return match.group(1), match.group(2)

    @staticmethod
    def _col_to_index(col: str) -> int:
        col = col.upper()
        index = 0

        for ch in col:
            index = index * 26 + (ord(ch) - ord('A') + 1)

        return index

    # -------------------- Cell Value Helpers -------------------- #
    @staticmethod
    def _remove_existing_value_nodes(cell_node: et._Element) -> None:
        """Remove existing value nodes (v, is, f) from cell."""
        for child in list(cell_node):
            tag_local = child.tag.split("}")[-1]
            if tag_local in {"v", "is", "f"}:
                cell_node.remove(child)

    @staticmethod
    def _create_numeric_cell_value(cell_node: et._Element, value: float, is_percent_style: bool) -> None:
        """
        Create numeric value node with percent handling.
        Keeps integer formatting when possible.
        """
        value_to_write = value / 100.0 if is_percent_style else value
        # ensure numeric text written sensibly
        if float(value_to_write).is_integer():
            text = str(int(float(value_to_write)))
        else:
            text = repr(float(value_to_write))

        v_node = et.SubElement(cell_node, f"{{{MAIN_NS}}}v")
        v_node.text = text
        cell_node.attrib.pop("t", None)  # ensure numeric
        logger.debug(f"Writing numeric {v_node.text} to {cell_node.attrib.get('r')}, percent_style={is_percent_style}")

    def _create_text_cell_value(self, cell_node: et._Element, value: Optional[Union[str, bool]]) -> None:
        """Create text cell value using shared strings."""
        text = "" if value is None else str(value)
        string_index = self.shared_strings.get_or_add(text)
        cell_node.set("t", "s")  # shared string
        v_node = et.SubElement(cell_node, f"{{{MAIN_NS}}}v")
        v_node.text = str(string_index)
        logger.debug(f"Writing text '{text}' to {cell_node.attrib.get('r')} as shared string index {string_index}")

    # -------------------- Ensure cell exists (ordered insertion) -------------------- #
    def ensure_cell_exists(self, root: et._Element, cell_ref: str, nsmap: Dict[str, str]) -> et._Element:
        """Ensure cell exists; insert row/cell preserving ascending order."""
        cell_node = self.find_ns(root, f".//main:c[@r='{cell_ref}']", nsmap)
        if cell_node is not None:
            return cell_node

        # Validate cell reference format
        col_letters, row_idx = self._parse_cell_ref(cell_ref)

        # Find or create sheetData
        sheet_data = self.find_ns(root, ".//main:sheetData", nsmap)

        if sheet_data is None:
            sheet_data = et.SubElement(root, f"{{{MAIN_NS}}}sheetData")

        # Find or insert row in order
        row_nodes = list(self.findall_ns(sheet_data, ".//main:row", nsmap))
        target_row_node = None
        inserted = False

        for i, rn in enumerate(row_nodes):
            r_attr = rn.attrib.get("r")

            try:
                r_val = int(r_attr) if r_attr is not None else None
            except ValueError:
                r_val = None

            if r_val is not None and r_val == int(row_idx):
                target_row_node = rn
                break

            if r_val is not None and r_val > int(row_idx):
                row_attrs = {"r": row_idx}

                if any("x14ac" in ns for ns in nsmap.values()):
                    row_attrs["{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent"] = "0.25"

                new_row = et.Element(f"{{{MAIN_NS}}}row", **row_attrs)
                sheet_data.insert(i, new_row)
                target_row_node = new_row
                inserted = True
                break

        if target_row_node is None:
            if not inserted:
                # Append new row at end
                row_attrs = {"r": row_idx}
                if any("x14ac" in ns for ns in nsmap.values()):
                    row_attrs["{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent"] = "0.25"

                target_row_node = et.SubElement(sheet_data, f"{{{MAIN_NS}}}row", **row_attrs)

        # Ensure cell within row preserving column order
        existing_cells = list(self.findall_ns(target_row_node, ".//main:c", nsmap))
        insert_index = None

        for i, c in enumerate(existing_cells):
            r = c.attrib.get("r", "")
            col_existing_match = re.match(r"^([A-Za-z]+)", r)
            col_existing_letters = col_existing_match.group(1) if col_existing_match else ""

            if col_existing_letters and self._col_to_index(col_existing_letters) > self._col_to_index(col_letters):
                insert_index = i
                break

        if insert_index is not None:
            new_cell = et.Element(f"{{{MAIN_NS}}}c", r=cell_ref)
            target_row_node.insert(insert_index, new_cell)
            return new_cell
        else:
            return et.SubElement(target_row_node, f"{{{MAIN_NS}}}c", r=cell_ref)

    # -------------------- Update cell value (matches original behavior) -------------------- #
    def update_cell_value(self, cell_node: et._Element, value: Optional[Union[str, int, float, bool]]) -> None:
        """
        Update cell value, preserving style and using shared strings for text.

        Args:
            cell_node: The XML element representing the cell
            value: The value to write (None, string, number, or boolean)
        """
        # Clear existing value nodes: v, is, f
        self._remove_existing_value_nodes(cell_node)

        style_id = cell_node.attrib.get("s")
        is_percent_style = self._style_id_is_percent(style_id)

        # Handle numeric values (int/float) but not bool
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                val_float = float(value)
                if math.isnan(val_float) or math.isinf(val_float):
                    val_float = 0.0
            except (TypeError, ValueError):
                val_float = 0.0

            self._create_numeric_cell_value(cell_node, float(val_float), is_percent_style)
            return

        # Handle text and None via shared strings
        self._create_text_cell_value(cell_node, value)

    # -------------------- Workbook relationships -------------------- #
    def _build_sheet_name_to_file_map(self, xl_dir: Path) -> Dict[str, Path]:
        """Build mapping from sheet names to XML files using workbook relationships."""
        workbook_xml = xl_dir / "workbook.xml"
        if not workbook_xml.exists():
            raise FileNotFoundError(f"Workbook XML not found: {workbook_xml}")

        tree = et.parse(workbook_xml, XML_PARSER)
        root = tree.getroot()
        nsmap = self.get_nsmap_with_main(root)
        sheets = self.findall_ns(root, ".//main:sheets/main:sheet", nsmap)

        # Parse relationships
        rels_map: Dict[str, str] = {}
        rels_path = xl_dir.parent / XL_RELS_PATH

        if rels_path.exists():
            try:
                rels_tree = et.parse(rels_path, XML_PARSER)
                rels_root = rels_tree.getroot()

                for rel in rels_root:
                    if rel.tag.endswith("}Relationship"):
                        rid = rel.attrib.get("Id")
                        target = rel.attrib.get("Target")

                        if rid and target:
                            # Normalize target path
                            if target.startswith("/"):
                                target = target[1:]
                            rels_map[rid] = target
            except Exception as ex:
                logger.warning(f"Failed parsing workbook relationships: {ex}")

        mapping: Dict[str, Path] = {}
        for sheet in sheets:
            name = sheet.attrib.get("name")
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")

            if not name:
                continue

            target_path = None
            if rid and rid in rels_map:
                rel_target = rels_map[rid]
                candidate = xl_dir.parent / rel_target

                if candidate.exists():
                    target_path = candidate
                else:
                    # Try relative to xl directory
                    candidate = xl_dir / rel_target
                    if candidate.exists():
                        target_path = candidate

            if target_path is None:
                # Fallback to default naming convention
                target_path = xl_dir / self._WORKSHEETS_DIR_NAME / f"sheet{len(mapping) + 1}.xml"

            if target_path.exists():
                mapping[name] = target_path
            else:
                logger.warning(f"Sheet file not found: {target_path}")

        return mapping

    # -------------------- Main XLSX updater (static facade preserved) -------------------- #
    def _update_xlsx_in_place(self, source: Path, target: Path, mappings: List[Dict[str, str]]) -> None:
        """Instance worker implementing original semantics, using instance-level caches."""
        # Validate inputs
        self._validate_inputs(source, target, mappings)

        if not mappings:
            return

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            logger.info(f"Processing in temporary directory: {temp_path}")

            # Working copy
            temp_copy = temp_path / self._TEMP_WORKBOOK_NAME
            shutil.copy2(target, temp_copy)

            # Unzip
            content_dir = temp_path / self._CONTENT_DIR_NAME
            try:
                with zipfile.ZipFile(temp_copy, "r") as zip_in:
                    zip_in.extractall(content_dir)
            except zipfile.BadZipFile as e:
                logger.error(f"Invalid ZIP archive: {e}")
                raise ValueError("File is not a valid Excel file") from e

            xl_dir = content_dir / "xl"

            # Load styles and shared strings
            self._load_styles_from_xl(xl_dir)
            self.shared_strings.load(xl_dir)
            self.shared_strings._modified = False  # reset modified flag after load

            # Extract source values
            src_values = ExcelUpdater.extract_source_values(source, mappings)
            if not src_values:
                logger.warning("No values extracted from source workbook")
                return

            # Map sheet names to XML files
            sheet_name_to_file = self._build_sheet_name_to_file_map(xl_dir)
            if not sheet_name_to_file:
                # Fallback
                sheets_dir = xl_dir / self._WORKSHEETS_DIR_NAME
                worksheet_files = list(sheets_dir.glob("sheet*.xml")) if sheets_dir.exists() else []

                if worksheet_files:
                    sheet_name_to_file = {self._DEFAULT_SHEET_NAME: worksheet_files[0]}
                else:
                    raise ValueError("No worksheets found in target workbook")

            # Organize updates by sheet file
            updates_by_sheet: Dict[Path, List[Tuple[str, Any]]] = {}
            default_sheet_file = next(iter(sheet_name_to_file.values()))

            for (sheet_name, cell_ref), value in src_values.items():
                sheet_file = sheet_name_to_file.get(sheet_name, default_sheet_file)
                updates_by_sheet.setdefault(sheet_file, []).append((cell_ref, value))

            # Apply updates
            total_updates = 0
            numeric_count = 0

            for sheet_file, updates in updates_by_sheet.items():
                try:
                    sheet_tree = et.parse(sheet_file, XML_PARSER)
                    sheet_root = sheet_tree.getroot()
                    nsmap = self.get_nsmap_with_main(sheet_root)

                    for cell_ref, value in updates:
                        cell_node = self.ensure_cell_exists(sheet_root, cell_ref, nsmap)
                        self.update_cell_value(cell_node, value)
                        total_updates += 1

                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            numeric_count += 1

                    # Write back sheet XML (no pretty print)
                    sheet_file.write_bytes(
                        et.tostring(sheet_root, xml_declaration=True, encoding="UTF-8",
                                    standalone="yes", pretty_print=False)
                    )
                except (et.ParseError, et.LxmlError) as e:
                    logger.error(f"XML parsing error in sheet {sheet_file}: {e}")
                    raise ValueError(f"Invalid Excel file structure: {e}") from e
                except Exception as ex:
                    logger.error(f"Failed updating sheet {sheet_file}: {ex}")
                    continue

            # Save shared strings if modified
            self.shared_strings.save(xl_dir)

            # Validate results
            self._validate_update_results(updates_by_sheet, total_updates)

            # Rezip contents
            updated_zip = temp_path / self._UPDATED_WORKBOOK_NAME
            with zipfile.ZipFile(updated_zip, "w", zipfile.ZIP_DEFLATED) as zip_out:
                for file_path in content_dir.rglob("*"):
                    if file_path.is_file():
                        zip_out.write(file_path, file_path.relative_to(content_dir))

            # Replace target
            shutil.copy2(updated_zip, target)

            # Log summary statistics
            text_count = total_updates - numeric_count
            logger.info(
                f"Successfully updated {total_updates} cells "
                f"({numeric_count} numeric, {text_count} text) "
                f"across {len(updates_by_sheet)} sheets"
            )

    # -------------------- Cleanup -------------------- #
    def cleanup(self) -> None:
        """Explicit cleanup method to clear large caches."""
        self._nsmap_cache.clear()
        self._style_to_numfmt.clear()

        if hasattr(self.shared_strings, '_text_to_index'):
            self.shared_strings._text_to_index.clear()

    # Static facade preserving original API
    @staticmethod
    def update_xlsx_in_place(source: Path, target: Path, mappings: List[Dict[str, str]]) -> None:
        updater = ExcelUpdater()

        try:
            return updater._update_xlsx_in_place(source, target, mappings)
        finally:
            updater.cleanup()
