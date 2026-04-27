"""
software_vn_parser.py
=====================
Parser for the **"Software VN"** Excel workbook (e.g. ``New software VN
20260420.xlsx``).  Sheet ``Info voor GPP`` shares the spirit of the
classic Verantwoordingsnota but the row layout is different (a new
"Virgin binder" row, three extra equipment-fuel rows, an Asfaltgranulaat
row, and a re-formatted plant address).

This parser is intentionally standalone — it does NOT delegate to
:mod:`vn_parser`, so the two upload flows can evolve independently.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, asdict, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Info voor GPP"

# Per-plant column letters: (general_value_col, transport_mode_col)
PLANT_COLS: dict[int, tuple[str, str]] = {
    0: ("E", "F"),
    1: ("G", "H"),
    2: ("I", "J"),
}

# ── Row anchors (Software VN layout, dated 2026-04-20) ──────────────
ROW_PLANT_HEADER       = 2
ROW_DATE               = 4
ROW_MIXTURE_ID         = 5
ROW_MIX_SB250          = 6
ROW_MIX_EN             = 7
ROW_BINDER_PCT_TOTAL   = 8
ROW_BINDER_REPL        = 9
ROW_VIRGIN_BINDER      = 10
ROW_PLANT_LOC          = 12
ROW_PLANT_ENERGY       = 13
ROW_PLANT_CAP          = 14
ROW_ENERGY_PRIM_SEC    = 15
ROW_PROD_TEMP          = 16
ROW_ELECTRIC_SHARE     = 17
ROW_WHEEL_LOADER_FUEL  = 18
ROW_BINDER_TYPE_HDR    = 20
ROW_BINDER_NAME        = 21
ROW_COMPOSITION_FIRST  = 26
ROW_COMPOSITION_LAST   = 51
ROW_BIOGENIC_PCT       = 52
ROW_ITSR               = 54
ROW_PRD                = 55
ROW_STIFF1             = 56
ROW_STIFF2             = 57
ROW_FATIGUE            = 58


# Origin placeholder strings that should be treated as "no origin".
_ORIGIN_PLACEHOLDERS = {"add location manually", "tbd", "n/a", "-"}


# ─────────────────────────────────────────────────────────────────────
#  Data classes
# ─────────────────────────────────────────────────────────────────────

@dataclass
class SVNComponent:
    row: int
    name: str
    pct: float
    origin: str | None
    mode: str | None
    extra: str | None = None


@dataclass
class SVNPlant:
    plant_index: int
    plant_id: str
    date: str | None = None
    mixture_id: str | None = None
    mixture_sb250: str | None = None
    mixture_en: str | None = None
    total_binder_pct: float | None = None
    binder_replacement_pct: float | None = None
    virgin_binder_pct: float | None = None
    plant_location: str | None = None
    plant_energy: str | None = None
    plant_capacity_tph: float | None = None
    energy_source_primary_secondary: str | None = None
    prod_temp_range: str | None = None
    electric_share_equipment: str | None = None
    wheel_loader_fuel: str | None = None
    binder_type: str | None = None
    binder_origin: str | None = None
    binder_mode: str | None = None
    binder_pct: float | None = None
    components: list[SVNComponent] = field(default_factory=list)
    biogenic_pct: float | None = None
    itsr: float | None = None
    prd: float | None = None
    stiffness_e_modulus: float | None = None
    fatigue_eps6: float | None = None

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["components"] = [asdict(c) for c in self.components]
        return d


@dataclass
class SoftwareVNData:
    source_filename: str | None
    plants: list[SVNPlant] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_filename": self.source_filename,
            "plants": [p.to_dict() for p in self.plants],
        }


# ─────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────

def _val(ws, row: int, col: str) -> Any:
    v = ws[f"{col}{row}"].value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _clean_origin(raw: Any) -> str | None:
    s = _to_str(raw)
    if s is None:
        return None
    if s.lower() in _ORIGIN_PLACEHOLDERS:
        return None
    return s


def _normalize_address(loc: str | None) -> str | None:
    """Re-arrange ``city ,zip ,street`` → ``street, zip city`` so the
    geocoder gets a normal-looking address."""
    if not loc:
        return loc
    parts = [p.strip() for p in loc.split(",") if p.strip()]
    if len(parts) == 3 and parts[1].isdigit() and len(parts[1]) == 4:
        city, zipc, street = parts
        return f"{street}, {zipc} {city}"
    return loc


# ─────────────────────────────────────────────────────────────────────
#  Public parser
# ─────────────────────────────────────────────────────────────────────

def parse(source: str | Path | bytes | io.BytesIO,
          source_filename: str | None = None) -> SoftwareVNData:
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    elif isinstance(source, io.BytesIO):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)
        if source_filename is None:
            source_filename = Path(source).name

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    plants: list[SVNPlant] = []
    for idx in range(3):
        gen_col, mode_col = PLANT_COLS[idx]
        plant_id = _to_str(_val(ws, ROW_PLANT_HEADER, gen_col)) or f"plant_{idx + 1}"

        components: list[SVNComponent] = []
        for row in range(ROW_COMPOSITION_FIRST, ROW_COMPOSITION_LAST + 1):
            name = _to_str(_val(ws, row, "C"))
            pct_raw = _val(ws, row, "D")
            pct = _to_float(pct_raw)
            extra = pct_raw if (pct is None and pct_raw is not None) else None
            origin = _clean_origin(_val(ws, row, gen_col))
            mode = _to_str(_val(ws, row, mode_col))

            if not name:
                continue
            if name.lower() in {"additieven", "aggregaten"}:
                continue
            if (pct is None or pct == 0) and not extra:
                continue

            components.append(SVNComponent(
                row=row,
                name=name,
                pct=pct if pct is not None else 0.0,
                origin=origin,
                mode=mode,
                extra=_to_str(extra),
            ))

        plant = SVNPlant(
            plant_index=idx,
            plant_id=plant_id,
            date=_to_str(_val(ws, ROW_DATE, gen_col)),
            mixture_id=_to_str(_val(ws, ROW_MIXTURE_ID, gen_col)),
            mixture_sb250=_to_str(_val(ws, ROW_MIX_SB250, gen_col)),
            mixture_en=_to_str(_val(ws, ROW_MIX_EN, gen_col)),
            total_binder_pct=_to_float(_val(ws, ROW_BINDER_PCT_TOTAL, gen_col)),
            binder_replacement_pct=_to_float(_val(ws, ROW_BINDER_REPL, gen_col)),
            virgin_binder_pct=_to_float(_val(ws, ROW_VIRGIN_BINDER, gen_col)),
            plant_location=_normalize_address(_to_str(_val(ws, ROW_PLANT_LOC, gen_col))),
            plant_energy=_to_str(_val(ws, ROW_PLANT_ENERGY, gen_col)),
            plant_capacity_tph=_to_float(_val(ws, ROW_PLANT_CAP, gen_col)),
            energy_source_primary_secondary=_to_str(_val(ws, ROW_ENERGY_PRIM_SEC, gen_col)),
            prod_temp_range=_to_str(_val(ws, ROW_PROD_TEMP, gen_col)),
            electric_share_equipment=_to_str(_val(ws, ROW_ELECTRIC_SHARE, gen_col)),
            wheel_loader_fuel=_to_str(_val(ws, ROW_WHEEL_LOADER_FUEL, gen_col)),
            binder_type=_to_str(_val(ws, ROW_BINDER_NAME, "C")),
            binder_origin=_clean_origin(_val(ws, ROW_BINDER_NAME, gen_col)),
            binder_mode=_to_str(_val(ws, ROW_BINDER_NAME, mode_col)),
            # For "synthetisch pigmenteerbaar bindmiddel" the GPP-tool
            # expects the *virgin* binder %, not the total binder %.
            # Software VN row 10 (Virgin binder) is authoritative; fall
            # back to column D of the binder row only if row 10 is empty.
            binder_pct=(
                _to_float(_val(ws, ROW_VIRGIN_BINDER, gen_col))
                or _to_float(_val(ws, ROW_BINDER_NAME, "D"))
            ),
            components=components,
            biogenic_pct=_to_float(_val(ws, ROW_BIOGENIC_PCT, gen_col)),
            itsr=_to_float(_val(ws, ROW_ITSR, gen_col)),
            prd=_to_float(_val(ws, ROW_PRD, gen_col)),
            stiffness_e_modulus=_to_float(_val(ws, ROW_STIFF1, gen_col)),
            fatigue_eps6=_to_float(_val(ws, ROW_FATIGUE, gen_col)),
        )
        plants.append(plant)

    return SoftwareVNData(source_filename=source_filename, plants=plants)
