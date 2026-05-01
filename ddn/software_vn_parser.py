"""
software_vn_parser.py
=====================
Parser for the **"Software VN"** Excel workbook (e.g. ``New software VN
20260420.xlsx``).  Sheet ``Info voor GPP`` shares the spirit of the
classic Verantwoordingsnota but the row layout is different (a new
"Virgin binder" row, three extra equipment-fuel rows, an Asfaltgranulaat
row, and a re-formatted plant address).

This parser is intentionally standalone — it does NOT delegate to
:mod:`vn_parser`, so the two upload flows can evolve independently.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, asdict, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Info voor GPP"

# Per-plant column letters: (general_value_col, transport_mode_col)
PLANT_COLS: dict[int, tuple[str, str]] = {
    0: ("E", "F"),
    1: ("G", "H"),
    2: ("I", "J"),
}

# ── Row anchors (Software VN layout, dated 2026-04-28b) ─────────────
# Layout in the wild (April 2026) keeps the "Virgin binder" row at 10
# and pushes the rest of section 2/3 down accordingly.  An earlier
# experimental Desktop variant removed row 10 — we don't support that
# variant; users must keep the Virgin binder row.
ROW_PLANT_HEADER       = 2
ROW_DATE               = 4
ROW_MIXTURE_ID         = 5
ROW_MIX_SB250          = 6
ROW_MIX_EN             = 7
ROW_BINDER_PCT_TOTAL   = 8
ROW_BINDER_REPL        = 9
ROW_VIRGIN_BINDER      = 10
ROW_PLANT_LOC          = 12
ROW_PLANT_ENERGY       = 13   # primary heater type (aardgas / propaan / elektrisch)
ROW_PRIM_ENERGY_PCT    = 14   # primary heater % share (0-1 fraction)
ROW_ENERGY_PRIM_SEC    = 15   # secondary heater type
ROW_SEC_ENERGY_PCT     = 16   # secondary heater % share (0-1 fraction)
ROW_PLANT_CAP          = 17
ROW_PROD_TEMP          = 18
ROW_ELECTRIC_SHARE     = 19   # "Ja"/"Nee" — are equipment electric?
ROW_ELECTRIC_SOURCE    = 20   # source for the electric equipment, e.g. "Elektriciteit normaal"
ROW_WHEEL_LOADER_FUEL  = 21
ROW_BINDER_TYPE_HDR    = 23
ROW_BINDER_NAME        = 24
ROW_COMPOSITION_FIRST  = 29
ROW_COMPOSITION_LAST   = 55
ROW_COMPOSITION_TOTAL  = 66   # "Total ( excl. Rode Kleurstof, Trinidad, Uintah)" in column D
ROW_BIOGENIC_PCT       = 56
ROW_ITSR               = 58
ROW_PRD                = 59
ROW_STIFF1             = 60
ROW_STIFF2             = 61
ROW_FATIGUE            = 62


# Origin placeholder strings that should be treated as "no origin".
_ORIGIN_PLACEHOLDERS = {"add location manually", "tbd", "n/a", "-"}


# ─────────────────────────────────────────────────────────────────────
#  Data classes
# ─────────────────────────────────────────────────────────────────────

@dataclass
class SVNComponent:
    row: int
    name: str
    pct: float
    origin: str | None
    mode: str | None
    extra: str | None = None


@dataclass
class SVNPlant:
    plant_index: int
    plant_id: str
    date: str | None = None
    mixture_id: str | None = None
    mixture_sb250: str | None = None
    mixture_en: str | None = None
    total_binder_pct: float | None = None
    binder_replacement_pct: float | None = None
    virgin_binder_pct: float | None = None
    plant_location: str | None = None
    plant_energy: str | None = None
    plant_capacity_tph: float | None = None
    primary_energy_pct: float | None = None
    energy_source_primary_secondary: str | None = None
    secondary_energy_pct: float | None = None
    prod_temp_range: str | None = None
    electric_share_equipment: str | None = None
    electric_source: str | None = None
    wheel_loader_fuel: str | None = None
    binder_type: str | None = None
    binder_origin: str | None = None
    binder_mode: str | None = None
    binder_pct: float | None = None
    components: list[SVNComponent] = field(default_factory=list)
    biogenic_pct: float | None = None
    composition_total_pct: float | None = None  # value of the VN's own "Total" cell (D66)
    itsr: float | None = None
    prd: float | None = None
    stiffness_e_modulus: float | None = None
    fatigue_eps6: float | None = None

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["components"] = [asdict(c) for c in self.components]
        return d


@dataclass
class SoftwareVNData:
    source_filename: str | None
    plants: list[SVNPlant] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_filename": self.source_filename,
            "plants": [p.to_dict() for p in self.plants],
        }


# ─────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────

def _val(ws, row: int, col: str) -> Any:
    v = ws[f"{col}{row}"].value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_pct(v: Any) -> float | None:
    """Normalize a cell to a 0–100 percentage.

    Excel cells formatted as Percent return their underlying fraction
    (e.g. 5.5% → 0.055), while plain-number cells return 5.5. We assume
    any value <= 1.0 was a fraction and scale it; values > 1 are taken
    as already-percent. The 0–1 ambiguity at exactly 1.0 is treated as
    1% (the safer assumption for binder / biogenic shares which are
    rarely 100%).
    """
    f = _to_float(v)
    if f is None:
        return None
    if -1.0 <= f <= 1.0:
        return f * 100.0
    return f


def _to_fraction(v: Any) -> float | None:
    """Normalize a cell to a 0–1 fraction (inverse of `_to_pct`)."""
    f = _to_float(v)
    if f is None:
        return None
    if f > 1.0:
        return f / 100.0
    return f


def _clean_origin(raw: Any) -> str | None:
    s = _to_str(raw)
    if s is None:
        return None
    if s.lower() in _ORIGIN_PLACEHOLDERS:
        return None
    return s


# Common yes/no spellings encountered in Software VN files. The
# canonical Dutch values are "Ja" / "Nee" — normalize English and
# truthy/falsey variants so downstream code (and the UI) sees a
# consistent value.
_YES_VALUES = {"ja", "yes", "y", "true", "1", "x"}
_NO_VALUES  = {"nee", "neen", "no", "n", "false", "0", "geen", "none"}


def _normalize_yes_no(raw: Any) -> str | None:
    s = _to_str(raw)
    if s is None:
        return None
    key = s.lower()
    if key in _YES_VALUES:
        return "Ja"
    if key in _NO_VALUES:
        return "Nee"
    return s


def _normalize_address(loc: str | None) -> str | None:
    """Re-arrange ``city ,zip ,street`` → ``street, zip city`` so the
    geocoder gets a normal-looking address. Handles extra whitespace,
    optional country suffix, and Belgian/Dutch 4-digit postcodes.
    """
    if not loc:
        return loc
    s = " ".join(loc.split())  # collapse whitespace
    parts = [p.strip() for p in s.split(",") if p.strip()]
    # Look for a 4-digit postcode token; if it's the middle of a 3-part
    # `city, zip, street` shape, flip it.
    if len(parts) == 3 and parts[1].isdigit() and len(parts[1]) == 4:
        city, zipc, street = parts
        return f"{street}, {zipc} {city}"
    # 4-part with trailing country: `city, zip, street, country`
    if len(parts) == 4 and parts[1].isdigit() and len(parts[1]) == 4:
        city, zipc, street, country = parts
        return f"{street}, {zipc} {city}, {country}"
    return s


# ─────────────────────────────────────────────────────────────────────
#  Public parser
# ─────────────────────────────────────────────────────────────────────

def parse(source: str | Path | bytes | io.BytesIO,
          source_filename: str | None = None) -> SoftwareVNData:
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    elif isinstance(source, io.BytesIO):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)
        if source_filename is None:
            source_filename = Path(source).name

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    plants: list[SVNPlant] = []
    # Dynamically locate the "Total" row instead of hardcoding D66 —
    # template tweaks (added/removed rows in the composition block)
    # would otherwise silently misread.
    composition_total_row: int | None = None
    for r in range(ROW_COMPOSITION_LAST + 1, ROW_COMPOSITION_LAST + 25):
        label = _to_str(_val(ws, r, "C")) or _to_str(_val(ws, r, "B"))
        if label and label.lower().lstrip().startswith("total"):
            composition_total_row = r
            break
    if composition_total_row is None:
        composition_total_row = ROW_COMPOSITION_TOTAL

    for idx in range(3):
        gen_col, mode_col = PLANT_COLS[idx]
        plant_id = _to_str(_val(ws, ROW_PLANT_HEADER, gen_col)) or f"plant_{idx + 1}"

        components: list[SVNComponent] = []
        for row in range(ROW_COMPOSITION_FIRST, ROW_COMPOSITION_LAST + 1):
            name = _to_str(_val(ws, row, "C"))
            pct_raw = _val(ws, row, "D")
            pct = _to_float(pct_raw)
            extra = pct_raw if (pct is None and pct_raw is not None) else None
            origin = _clean_origin(_val(ws, row, gen_col))
            mode = _to_str(_val(ws, row, mode_col))

            if not name:
                continue
            if name.lower() in {"additieven", "aggregaten"}:
                continue
            # Drop a row only when there is genuinely nothing to record:
            # no %, no extra annotation, AND no operator-entered
            # origin/mode metadata.
            if (pct is None or pct == 0) and not extra and not origin and not mode:
                continue

            components.append(SVNComponent(
                row=row,
                name=name,
                pct=pct if pct is not None else 0.0,
                origin=origin,
                mode=mode,
                extra=_to_str(extra),
            ))

        plant = SVNPlant(
            plant_index=idx,
            plant_id=plant_id,
            date=_to_str(_val(ws, ROW_DATE, gen_col)),
            mixture_id=_to_str(_val(ws, ROW_MIXTURE_ID, gen_col)),
            mixture_sb250=_to_str(_val(ws, ROW_MIX_SB250, gen_col)),
            mixture_en=_to_str(_val(ws, ROW_MIX_EN, gen_col)),
            total_binder_pct=_to_pct(_val(ws, ROW_BINDER_PCT_TOTAL, gen_col)),
            binder_replacement_pct=_to_pct(_val(ws, ROW_BINDER_REPL, gen_col)),
            virgin_binder_pct=_to_pct(_val(ws, ROW_VIRGIN_BINDER, gen_col)),
            plant_location=_normalize_address(_to_str(_val(ws, ROW_PLANT_LOC, gen_col))),
            plant_energy=_to_str(_val(ws, ROW_PLANT_ENERGY, gen_col)),
            plant_capacity_tph=_to_float(_val(ws, ROW_PLANT_CAP, gen_col)),
            primary_energy_pct=_to_float(_val(ws, ROW_PRIM_ENERGY_PCT, gen_col)),
            energy_source_primary_secondary=_to_str(_val(ws, ROW_ENERGY_PRIM_SEC, gen_col)),
            secondary_energy_pct=_to_float(_val(ws, ROW_SEC_ENERGY_PCT, gen_col)),
            prod_temp_range=_to_str(_val(ws, ROW_PROD_TEMP, gen_col)),
            electric_share_equipment=_normalize_yes_no(_val(ws, ROW_ELECTRIC_SHARE, gen_col)),
            electric_source=_to_str(_val(ws, ROW_ELECTRIC_SOURCE, gen_col)),
            wheel_loader_fuel=_to_str(_val(ws, ROW_WHEEL_LOADER_FUEL, gen_col)),
            binder_type=_to_str(_val(ws, ROW_BINDER_NAME, "C")),
            binder_origin=_clean_origin(_val(ws, ROW_BINDER_NAME, gen_col)),
            binder_mode=_to_str(_val(ws, ROW_BINDER_NAME, mode_col)),
            # For "synthetisch pigmenteerbaar bindmiddel" the GPP-tool
            # expects the *virgin* binder %, not the total binder %.
            # Software VN row 10 (Virgin binder) is authoritative; fall
            # back to column D of the binder row only when row 10 is
            # truly empty (None) — do NOT fall through on a legitimate
            # 0.0 (e.g. fully recycled mix), which `or` would silently do.
            binder_pct=(
                _to_pct(_val(ws, ROW_VIRGIN_BINDER, gen_col))
                if _val(ws, ROW_VIRGIN_BINDER, gen_col) is not None
                else _to_pct(_val(ws, ROW_BINDER_NAME, "D"))
            ),
            components=components,
            biogenic_pct=_to_pct(_val(ws, ROW_BIOGENIC_PCT, gen_col)),
            composition_total_pct=_to_pct(_val(ws, composition_total_row, "D")),
            itsr=_to_float(_val(ws, ROW_ITSR, gen_col)),
            prd=_to_float(_val(ws, ROW_PRD, gen_col)),
            stiffness_e_modulus=_to_float(_val(ws, ROW_STIFF1, gen_col)),
            fatigue_eps6=_to_float(_val(ws, ROW_FATIGUE, gen_col)),
        )
        plants.append(plant)

    return SoftwareVNData(source_filename=source_filename, plants=plants)
